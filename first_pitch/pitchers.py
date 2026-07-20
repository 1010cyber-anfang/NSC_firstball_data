
import openpyxl
from collections import defaultdict
import sys
sys.stdout.reconfigure(encoding="utf-8")

FILE = "2026垒球锦标赛.xlsx"


def norm(v):
    if v is None:
        return None
    t = str(v).strip().upper().replace("　", " ")
    t = " ".join(t.split())
    t = t.replace("1 SH", "1SH").replace("1 FC", "1FC").replace("1 FO", "1FO")
    if t in ("", "NA", "IBB"):
        return None
    return t

SWING  = {"1S","1F","1FO","1FC","1SH","1E","1T","1HR","1IB","1 1B","1 2B","1 3B"}
STRIKE = SWING | {"0S"}
ONBASE = {"1 1B","1 2B","1 3B","1HR","1FC"}
INPLAY = {"1F","1FO","1FC","1SH","1E","1T","1HR","1IB","1 1B","1 2B","1 3B"}
VALID  = SWING | {"0S","0B","HBP"}

def load(path=FILE):
    ws = openpyxl.load_workbook(path)["Sheet1"]
    col_team, cur = {}, None
    for c in range(3, ws.max_column+1):
        v = ws.cell(1,c).value
        if v not in (None,""): cur = v
        col_team[c] = cur
    pit = {c:(col_team[c], str(ws.cell(2,c).value).strip()) for c in range(3, ws.max_column+1)}
    rows, cur = [], None
    for r in range(3, ws.max_row+1):
        v = ws.cell(r,1).value
        if v not in (None,""): cur = v
        if ws.cell(r,2).value not in (None,""):
            rows.append((r, (cur, str(ws.cell(r,2).value).strip())))
    events = []
    for r, bat in rows:
        for c in range(3, ws.max_column+1):
            t = norm(ws.cell(r,c).value)
            if t is not None:
                if t not in VALID:
                    raise ValueError(f"unknown token {t!r} at row {r} col {c}")
                events.append((bat, pit[c], t))
    return events

pct = lambda a,b: f"{a/b:.1%}" if b else "-"

MIN_N = 30
agg = defaultdict(lambda: defaultdict(int))
for bat, pit, t in load():
    d = agg[pit]
    d["n"] += 1
    d["strike"] += t in STRIKE
    d["onbase"] += t in ONBASE

rows = [(p, d) for p, d in agg.items() if d["n"] >= MIN_N]
rows.sort(key=lambda x: -x[1]["strike"] / x[1]["n"])
print(f"{'Team':<6} {'Pitcher':<10} {'n':>4} {'Strike%':>9} {'OB%':>7} {'OB':>4}")
print("-" * 46)
for (team, name), d in rows:
    print(f"{team:<6} {name:<10} {d['n']:>4} {pct(d['strike'], d['n']):>9} "
          f"{pct(d['onbase'], d['n']):>7} {d['onbase']:>4}")