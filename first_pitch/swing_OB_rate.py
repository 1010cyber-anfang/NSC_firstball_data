
import openpyxl
from collections import defaultdict
import sys
sys.stdout.reconfigure(encoding="utf-8")
FILE = "2026垒球锦标赛.xlsx"
def norm(v):
    if v is None:
        return None
    t = str(v).strip().upper().replace("　", " ")
    t = " ".join(t.split())
    t = t.replace("1 SH", "1SH").replace("1 FC", "1FC").replace("1 FO", "1FO")
    if t in ("", "NA", "IBB"):
        return None
    return t

SWING  = {"1S","1F","1FO","1FC","1SH","1E","1T","1HR","1IB","1 1B","1 2B","1 3B"}
STRIKE = SWING | {"0S"}
ONBASE = {"1 1B","1 2B","1 3B","1HR","1FC"}
INPLAY = {"1F","1FO","1FC","1SH","1E","1T","1HR","1IB","1 1B","1 2B","1 3B"}
VALID  = SWING | {"0S","0B","HBP"}

def load(path=FILE):
    ws = openpyxl.load_workbook(path)["Sheet1"]
    col_team, cur = {}, None
    for c in range(3, ws.max_column+1):
        v = ws.cell(1,c).value
        if v not in (None,""): cur = v
        col_team[c] = cur
    pit = {c:(col_team[c], str(ws.cell(2,c).value).strip()) for c in range(3, ws.max_column+1)}
    rows, cur = [], None
    for r in range(3, ws.max_row+1):
        v = ws.cell(r,1).value
        if v not in (None,""): cur = v
        if ws.cell(r,2).value not in (None,""):
            rows.append((r, (cur, str(ws.cell(r,2).value).strip())))
    events = []
    for r, bat in rows:
        for c in range(3, ws.max_column+1):
            t = norm(ws.cell(r,c).value)
            if t is not None:
                if t not in VALID:
                    raise ValueError(f"unknown token {t!r} at row {r} col {c}")
                events.append((bat, pit[c], t))
    return events

pct = lambda a,b: f"{a/b:.1%}" if b else "-"
import math

MIN_N = 10
HITS = {"1 1B", "1 2B", "1 3B", "1HR", "1T"}

agg = defaultdict(lambda: defaultdict(int))
for bat, pit, t in load():
    d = agg[bat]
    d["n"] += 1
    d["swing"] += t in SWING
    d["hit"]   += t in HITS

pts = [(d["swing"] / d["n"], d["hit"] / d["n"], b, d)
       for b, d in agg.items() if d["n"] >= MIN_N]

xs = [p[0] for p in pts]
ys = [p[1] for p in pts]
n  = len(pts)
mx, my = sum(xs) / n, sum(ys) / n
sxx = sum((x - mx) ** 2 for x in xs)
syy = sum((y - my) ** 2 for y in ys)
sxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
r = sxy / math.sqrt(sxx * syy)
tstat = r * math.sqrt((n - 2) / (1 - r * r))
p = math.erfc(abs(tstat) / math.sqrt(2))

print(f"Batters (n >= {MIN_N})            : {n}")
print(f"Pearson r (swing% vs hit%/PA)  : {r:+.3f}")
print(f"t = {tstat:.2f}, two-sided p ~ {p:.3f}")
print()

pts.sort(key=lambda p: p[0])
k = n // 3
groups = [("Low swing%", pts[:k]), ("Mid swing%", pts[k:2*k]), ("High swing%", pts[2*k:])]
print(f"{'Group':<12} {'batters':>7} {'avg swing%':>11} {'hit rate':>9} {'hits/PA':>12}")
print("-" * 56)
for name, g in groups:
    npa    = sum(d["n"] for *_, d in g)
    nhit   = sum(d["hit"] for *_, d in g)
    avg_sw = sum(x for x, *_ in g) / len(g)
    print(f"{name:<12} {len(g):>7} {avg_sw:>11.1%} {pct(nhit, npa):>9} {nhit:>5}/{npa}")

print()
print(f"{'Team':<6} {'Batter':<12} {'n':>4} {'Swing%':>8} {'Hit%':>7} {'Hits':>5}")
print("-" * 48)
for x, y, (team, name), d in sorted(pts, key=lambda p: -p[1]):
    print(f"{team:<6} {name:<12} {d['n']:>4} {x:>8.1%} {y:>7.1%} {d['hit']:>5}")